"""OddsFlow V4 — consolidation workbook builder (data-embedded).

Built from: the foundation backtest (app/engine/static_policy.V3_MARKETS, the
2026-05-30 from-scratch test on 28,571 settled fixtures) + LIVE measurements
computed this session from the running database (tier x market, signal
performance, Picks Log filter lift, live-emit performance). Country/region and a
fresh full-DB recompute are pending the 300 MB DB read (sandbox mount truncates
it); see the Notes sheet.
"""
import os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "analysis")
os.makedirs(OUT, exist_ok=True)
STAMP = "2026-06-23"
PATH = os.path.join(OUT, f"OddsFlow_V4_Consolidation_{STAMP}.xlsx")

# ---------------- data ----------------
# Foundation matrix (V3_MARKETS): (zone,bts) -> dict
FOUND = {
    ("strong","over"):    dict(g=71.7,gn=2941,c=69.8,cn=2281,cl=7.5,t=69.4,tn=2941,comp=70.3),
    ("strong","under"):   dict(g=66.7,gn=4815,c=67.6,cn=3259,cl=7.5,t=74.3,tn=4815,comp=69.5),
    ("standard","over"):  dict(g=77.3,gn=10996,c=64.0,cn=8382,cl=8.5,t=72.7,tn=10996,comp=71.3),
    ("standard","under"): dict(g=71.8,gn=2062,c=56.5,cn=1412,cl=8.5,t=81.5,tn=2062,comp=69.9),
    ("low","over"):       dict(g=80.8,gn=3147,c=66.2,cn=2740,cl=8.5,t=79.9,tn=3147,comp=75.6),
    ("low","under"):      dict(g=71.1,gn=802,c=55.8,cn=565,cl=8.5,t=88.4,tn=802,comp=71.8),
    ("one_sided","over"): dict(g=84.4,gn=2435,c=68.8,cn=2279,cl=8.5,t=88.1,tn=2435,comp=80.4),
    ("one_sided","under"):dict(g=82.3,gn=1373,c=65.1,cn=1175,cl=8.5,t=94.3,tn=1373,comp=80.6),
}
ZONES=["strong","standard","low","one_sided"]; BTS=["over","under"]

# Live tier x market (on-the-fly, all emits) — wins/settled
TIER={
 "T1":{"3W":(92,121),"G":(134,181),"C":(102,141)},
 "T2":{"3W":(99,135),"G":(116,151),"C":(95,145)},
 "T3":{"3W":(109,150),"G":(153,178),"C":(105,163)},
}
# Signals (live) wins/settled per market
SPREAD={"strong":{"3W":(72,99),"G":(103,118),"C":(76,109)},
        "slight":{"3W":(228,307),"G":(300,392),"C":(226,340)}}
DF={"DF0":{"3W":(25,39),"G":(33,39),"C":(26,40)},
    "DF1":{"3W":(118,175),"G":(180,217),"C":(131,193)},
    "DF2":{"3W":(157,192),"G":(171,227),"C":(136,205)}}
# Picks Log filter (all settled) wins/settled
PL={"Qualifying (>=72% bar)":{"ALL":(341,414),"3W":(130,161),"G":(191,227),"C":(20,26)},
    "Filtered out (<72%)":{"ALL":(290,428),"3W":(89,127),"G":(37,53),"C":(164,248)},
    "All emits":{"ALL":(1005,1365),"3W":(300,406),"G":(403,510),"C":(302,449)}}
PL_WEEKEND={"Qualifying":(73,95),"Filtered out":(41,72)}

def pct(ws):
    w,n=ws; return round(100*w/n,1) if n else None
def frac(ws):
    w,n=ws; return f"{w}/{n}"

# ---------------- styling ----------------
wb=openpyxl.Workbook()
HEAD=Font(bold=True,color="FFFFFF"); HFILL=PatternFill("solid",fgColor="2F4858")
TITLE=Font(bold=True,size=14); SUB=Font(italic=True,color="666666"); B=Font(bold=True)
GREEN=PatternFill("solid",fgColor="C6EFCE"); AMBER=PatternFill("solid",fgColor="FFEB9C"); RED=PatternFill("solid",fgColor="FFC7CE")
CEN=Alignment(horizontal="center")
def fill(p):
    return None if p is None else (GREEN if p>=72 else (AMBER if p>=65 else RED))
def hrow(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.font=HEAD; x.fill=HFILL; x.alignment=CEN
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w

# Sheet 1 Overview
ws=wb.active; ws.title="Overview"
ws["A1"]="OddsFlow V4 — Consolidation"; ws["A1"].font=TITLE
ws["A2"]=f"{STAMP} · foundation backtest + live measurements (API subscription lapsed 2026-06-22)"; ws["A2"].font=SUB
r=4
land=[("Fixtures total","51,883"),("Fixtures settled","47,407 (≈91%)"),
      ("Emitted picks (settled/total)","1,521 / 1,891"),
      ("Engine hit rate — last 7d / 30d (events)","75.0% / 73.5%"),
      ("Active leagues","28 (pre-lapse)"),
      ("Foundation test base","28,571 settled fixtures (2026-05-30)")]
for k,v in land:
    ws.cell(row=r,column=1,value=k).font=B; ws.cell(row=r,column=2,value=v); r+=1
r+=1
ws.cell(row=r,column=1,value="What this workbook is").font=Font(bold=True,size=12); r+=1
for line in [
 "Two evidence layers: (1) FOUNDATION = the 2026-05-30 from-scratch backtest baked into the engine (large historical samples).",
 "(2) LIVE = picks the engine actually fired and how they settled, measured this session from the running DB.",
 "Partition = (draw_zone × BTS over/under) = 8 cells. Spread, DF, H2H are SIGNALS, not cell axes.",
 "Goals O1.5 = total≥2. Corners O7.5(strong)/O8.5 graded where stats exist. 3-Way = alpha-or-draw (binary).",
 "Hit rate = wins / settled. Green ≥72% · Amber 65–72% · Red <65%.",
 "PENDING (needs DB read, currently blocked by mount): country/region slice + a fresh full-DB recompute.",
]:
    ws.cell(row=r,column=1,value="• "+line); r+=1
widths(ws,[42,60])

# Sheet 2 Foundation Cell Matrix
ws=wb.create_sheet("Foundation Matrix")
ws.append(["FOUNDATION backtest — V3_MARKETS (2026-05-30, 28,571 fixtures)"]); ws["A1"].font=Font(bold=True,size=12)
hdr=["Zone","BTS","Goals %","n","Corners %","line","n","3-Way %","n","Composite %"]
ws.append(hdr); hrow(ws,2,len(hdr))
for z in ZONES:
    for b in BTS:
        d=FOUND[(z,b)]
        ws.append([z,b,d["g"],d["gn"],d["c"],d["cl"],d["cn"],d["t"],d["tn"],d["comp"]])
        rr=ws.max_row
        for col,p in ((3,d["g"]),(5,d["c"]),(8,d["t"]),(10,d["comp"])):
            f=fill(p)
            if f: ws.cell(row=rr,column=col).fill=f
ws.freeze_panes="A3"; widths(ws,[11,8,9,8,10,6,8,9,8,12])

# Sheet 3 Live by Tier
ws=wb.create_sheet("Live by Tier")
ws.append(["LIVE — emitted picks settled on the fly, by tier × market"]); ws["A1"].font=Font(bold=True,size=12)
ws.append(["Tier","3-Way %","3-Way","Goals %","Goals","Corners %","Corners","All %"]); hrow(ws,2,8)
for t in ("T1","T2","T3"):
    d=TIER[t]; tw=sum(v[0] for v in d.values()); tn=sum(v[1] for v in d.values())
    ws.append([t,pct(d["3W"]),frac(d["3W"]),pct(d["G"]),frac(d["G"]),pct(d["C"]),frac(d["C"]),round(100*tw/tn,1)])
    rr=ws.max_row
    for col,key in ((2,"3W"),(4,"G"),(6,"C")):
        f=fill(pct(d[key]))
        if f: ws.cell(row=rr,column=col).fill=f
    f=fill(round(100*tw/tn,1));
    if f: ws.cell(row=rr,column=8).fill=f
ws.freeze_panes="A3"; widths(ws,[8,9,9,9,9,10,9,8])

# Sheet 4 Signals
ws=wb.create_sheet("Signals")
ws.append(["LIVE signal performance — hit rate split by signal value, per market"]); ws["A1"].font=Font(bold=True,size=12)
ws.append(["Spread (BTS)","3-Way %","Goals %","Corners %"]); hrow(ws,2,4)
for sp in ("strong","slight"):
    d=SPREAD[sp]; ws.append([sp,pct(d["3W"]),pct(d["G"]),pct(d["C"])])
    rr=ws.max_row
    for col,key in ((2,"3W"),(3,"G"),(4,"C")):
        f=fill(pct(d[key]))
        if f: ws.cell(row=rr,column=col).fill=f
r=ws.max_row+2
ws.cell(row=r,column=1,value="DF (draw favouritism)").font=B
ws.append(["DF","3-Way %","Goals %","Corners %"]); hrow(ws,r+1,4)
for dfk in ("DF0","DF1","DF2"):
    d=DF[dfk]; ws.append([dfk,pct(d["3W"]),pct(d["G"]),pct(d["C"])])
    rr=ws.max_row
    for col,key in ((2,"3W"),(3,"G"),(4,"C")):
        f=fill(pct(d[key]))
        if f: ws.cell(row=rr,column=col).fill=f
r=ws.max_row+2
ws.cell(row=r,column=1,value="Read: spread=strong is a strong GOALS signal (+11pp); DF2 is a strong 3-WAY signal (+18pp vs DF0).").font=SUB
widths(ws,[18,10,10,11])

# Sheet 5 Picks Log Filter
ws=wb.create_sheet("Picks Log Filter")
ws.append(["LIVE — does the Picks Log 72% signal filter add value? (all settled)"]); ws["A1"].font=Font(bold=True,size=12)
ws.append(["Group","All %","All","3-Way %","Goals %","Corners %"]); hrow(ws,2,6)
for grp,d in PL.items():
    ws.append([grp,pct(d["ALL"]),frac(d["ALL"]),pct(d["3W"]),pct(d["G"]),pct(d["C"])])
    rr=ws.max_row
    f=fill(pct(d["ALL"]))
    if f: ws.cell(row=rr,column=2).fill=f
r=ws.max_row+2
ws.cell(row=r,column=1,value="Weekend (Sat–Sun) check:").font=B
ws.append(["Qualifying",pct(PL_WEEKEND["Qualifying"]),frac(PL_WEEKEND["Qualifying"])])
ws.append(["Filtered out",pct(PL_WEEKEND["Filtered out"]),frac(PL_WEEKEND["Filtered out"])])
r=ws.max_row+2
ws.cell(row=r,column=1,value="Read: the 72% bar lifts hit rate from 73.6% (all) to 82.4% (kept); filtered-out land 67.8%. ~15pp separation.").font=SUB
widths(ws,[24,9,9,10,10,11])

# Sheet 6 Rankings
ws=wb.create_sheet("Cell Rankings")
ws.append(["Foundation cells ranked by composite hit %"]); ws["A1"].font=Font(bold=True,size=12)
ws.append(["Rank","Zone","BTS","Composite %","Goals %","Corners %","3-Way %","n"]); hrow(ws,2,8)
ranked=sorted(FOUND.items(), key=lambda kv:-kv[1]["comp"])
for i,((z,b),d) in enumerate(ranked,1):
    ws.append([i,z,b,d["comp"],d["g"],d["c"],d["t"],d["tn"]])
    f=fill(d["comp"])
    if f: ws.cell(row=ws.max_row,column=4).fill=f
ws.freeze_panes="A3"; widths(ws,[6,11,8,13,9,10,9,8])

# Sheet 7 Notes
ws=wb.create_sheet("Notes & Pending")
ws["A1"]="Notes, data sources, and what still needs doing"; ws["A1"].font=Font(bold=True,size=12)
notes=[
 "DATA SOURCES:",
 "  Foundation Matrix = app/engine/static_policy.V3_MARKETS (2026-05-30 from-scratch test, 28,571 settled fixtures).",
 "  Live sheets = computed this session by re-settling the actual emit_log against final scores in the running DB.",
 "",
 "WHY TWO LAYERS DIFFER: foundation uses years of history per cell; live uses only the ~1,900 picks the engine has",
 "  fired since 2026-05-30. Live samples are smaller and noisier but reflect the post-overlay boundaries in production.",
 "",
 "PENDING (blocked this session, not forgotten):",
 "  • Country / region hit-rate slice — needs a full read of the 300 MB DB; the sandbox mount truncates it.",
 "  • Fresh full-DB recompute of the cell matrix (vs the frozen foundation) — same DB-read dependency.",
 "  • H2H-corner signal performance — not recorded in emit_log yet (only spread+DF are reconstructable).",
 "  Fix path: run scripts/consolidate.py on the host (reads the real DB directly) — it produces these plus country/region.",
 "",
 "HEADLINE READINGS:",
 "  • Goals O1.5 is the anchor market (live 79% blended; foundation 67–84% by cell).",
 "  • 3-Way (alpha-or-draw) strongest in one_sided & low-under cells (foundation 88–94%).",
 "  • Corners is the weakest leg, especially in under cells and lower tiers (~56–66%).",
 "  • Signals validated live: spread→Goals (+11pp), DF2→3-Way (+18pp). Picks Log 72% filter adds ~+9pp.",
]
for i,line in enumerate(notes,3):
    ws.cell(row=i,column=1,value=line)
widths(ws,[110])

wb.save(PATH)
print("WROTE", PATH)
