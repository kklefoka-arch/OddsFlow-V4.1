"""OddsFlow V4 — full-database consolidation workbook.

Runs ON THE HOST against the real data/oddsflow_v4.db (the sandbox mount can't
read the 300 MB DB reliably). Backtests EVERY settled+classifiable fixture and
writes a foundation-style multi-sheet .xlsx plus a short summary .txt.

Hit-rate convention (engine rules): goals_nl = O1.5 (total>=2); corners_nl =
O7.5 strong / O8.5 else (graded only where corner stats exist); threeway =
alpha-or-draw (favourite win OR draw = WIN, binary). No EV, no Wilson.
"""
from __future__ import annotations
import os, sqlite3, datetime, re

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(HERE, "data", "oddsflow_v4.db")
OUTDIR = os.path.join(HERE, "analysis")
os.makedirs(OUTDIR, exist_ok=True)
STAMP = datetime.date.today().isoformat()
XLSX = os.path.join(OUTDIR, f"OddsFlow_V4_Consolidation_{STAMP}.xlsx")
TXT = os.path.join(OUTDIR, f"consolidation_summary_{STAMP}.txt")

# ---------------- classification (inlined from app/engine/classify.py) ----------------
def zone_of(d):
    if d is None: return None
    if d < 2.90: return None
    if d < 3.30: return "strong"
    if d < 3.80: return "standard"
    if d < 4.30: return "low"
    return "one_sided"
def bts_yesno(y, n):
    if y is None or n is None: return None
    return "over" if y <= n else "under"
def bts_spread(y, n):
    if y is None or n is None: return None
    return "strong" if min(y, n) < 1.50 else "slight"
def df_of(ho, ao):
    if ho is None or ao is None: return None
    diff = abs(round(ho) - round(ao))
    return "DF0" if diff == 0 else ("DF1" if diff == 1 else "DF2")

ZONES = ["strong", "standard", "low", "one_sided"]
BTS = ["over", "under"]
MKTS = ["goals_nl", "corners_nl", "threeway"]
MLABEL = {"goals_nl": "Goals O1.5", "corners_nl": "Corners O7.5/8.5", "threeway": "3-Way (alpha-or-draw)"}

def goals_win(hs, aw):  # O1.5
    return 1 if (hs + aw) >= 2 else None if hs is None else (1 if (hs + aw) >= 2 else 0)
def corners_win(zone, tc):
    if tc is None: return None
    line = 7.5 if zone == "strong" else 8.5
    return 1 if tc > line else 0
def threeway_win(hs, aw, ho, ao):
    ah = (ho is None or ao is None or ho <= ao)
    if hs == aw: return 1
    return 1 if ((hs > aw) if ah else (aw > hs)) else 0

class Acc:
    __slots__ = ("g", "n")
    def __init__(self): self.g = 0; self.n = 0
    def add(self, h):
        if h is None: return
        self.g += h; self.n += 1
    def pct(self): return round(100 * self.g / self.n, 1) if self.n else None
    def frac(self): return f"{self.g}/{self.n}"

def newcell(): return {m: Acc() for m in MKTS}

# ---------------- load ----------------
print("Reading", DB)
conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
rows = conn.execute("""
  SELECT f.draw_odd, f.btts_yes_odd y, f.btts_no_odd n, f.home_odd ho, f.away_odd ao,
         f.home_score hs, f.away_score aw, fs.home_corners hc, fs.away_corners ac, fs.total_corners tc,
         lg.country country, lg.tier tier
  FROM fixtures f
  LEFT JOIN fixture_stats fs ON fs.fixture_id = f.id
  LEFT JOIN leagues lg ON lg.id = f.league_id
  WHERE f.home_score IS NOT NULL AND f.away_score IS NOT NULL AND f.draw_odd IS NOT NULL
    AND f.btts_yes_odd IS NOT NULL AND f.btts_no_odd IS NOT NULL
    AND f.home_odd IS NOT NULL AND f.away_odd IS NOT NULL
""").fetchall()
print(f"{len(rows)} settled+classifiable fixtures")

cells = {}        # (zone,bts) -> {market:Acc}
tier_mkt = {}     # (tier,market) -> Acc
country_mkt = {}  # (country,market) -> Acc
country_n = {}
spread_mkt = {}   # (spread,market) -> Acc
df_mkt = {}       # (df,market) -> Acc
corner_avail = {} # (zone,bts) -> count fixtures with corner data
def gp(d, k):
    if k not in d: d[k] = Acc()
    return d[k]

for r in rows:
    z = zone_of(r["draw_odd"]); b = bts_yesno(r["y"], r["n"])
    if z is None or b is None: continue
    tc = r["tc"]
    if tc is None and r["hc"] is not None and r["ac"] is not None:
        tc = r["hc"] + r["ac"]
    sp = bts_spread(r["y"], r["n"]); df = df_of(r["ho"], r["ao"])
    wins = {
        "goals_nl": 1 if (r["hs"] + r["aw"]) >= 2 else 0,
        "corners_nl": corners_win(z, tc),
        "threeway": threeway_win(r["hs"], r["aw"], r["ho"], r["ao"]),
    }
    if (z, b) not in cells: cells[(z, b)] = newcell(); corner_avail[(z, b)] = 0
    if tc is not None: corner_avail[(z, b)] += 1
    for m in MKTS:
        cells[(z, b)][m].add(wins[m])
        if r["tier"] in (1, 2, 3): gp(tier_mkt, (r["tier"], m)).add(wins[m])
        if r["country"]: gp(country_mkt, (r["country"], m)).add(wins[m]); country_n[r["country"]] = country_n.get(r["country"], 0) + (1 if m == "threeway" else 0)
        if sp: gp(spread_mkt, (sp, m)).add(wins[m])
        if df: gp(df_mkt, (df, m)).add(wins[m])

# ---------------- live emits (actual fired picks) ----------------
emit_rows = conn.execute("""
  SELECT em.market, em.zone, em.bts_pocket, em.df_level, em.tier, em.pick,
         f.home_score hs, f.away_score aw, f.home_odd ho, f.away_odd ao, fs.total_corners tc
  FROM emit_log em JOIN fixtures f ON f.id = em.fixture_id
  LEFT JOIN fixture_stats fs ON fs.fixture_id = f.id
  WHERE em.market IN ('goals_nl','corners_nl','threeway')
    AND f.home_score IS NOT NULL AND f.away_score IS NOT NULL
""").fetchall()
def settle_emit(m, z, hs, aw, ho, ao, pick, tc):
    if m == "goals_nl": return 1 if (hs + aw) >= 2 else 0
    if m == "corners_nl":
        if tc is None: return None
        mm = re.search(r"Over (\d+\.5)", pick or ""); line = float(mm.group(1)) if mm else (7.5 if z == "strong" else 8.5)
        return 1 if tc > line else 0
    if m == "threeway": return threeway_win(hs, aw, ho, ao)
emit_mkt = {m: Acc() for m in MKTS}
emit_tier = {}
for r in emit_rows:
    h = settle_emit(r["market"], r["zone"], r["hs"], r["aw"], r["ho"], r["ao"], r["pick"], r["tc"])
    emit_mkt[r["market"]].add(h)
    if r["tier"] in (1, 2, 3): gp(emit_tier, (r["tier"], r["market"])).add(h)

g1 = lambda q: conn.execute(q).fetchone()[0]
landscape = {
    "fixtures_total": g1("SELECT COUNT(*) FROM fixtures"),
    "settled": g1("SELECT COUNT(*) FROM fixtures WHERE home_score IS NOT NULL"),
    "classifiable_settled": len(rows),
    "corner_rows": g1("SELECT COUNT(*) FROM fixture_stats"),
    "leagues": g1("SELECT COUNT(*) FROM leagues"),
    "leagues_active": g1("SELECT COUNT(*) FROM leagues WHERE active=1"),
    "emit_log": g1("SELECT COUNT(*) FROM emit_log"),
    "date_min": conn.execute("SELECT MIN(date) FROM fixtures WHERE home_score IS NOT NULL").fetchone()[0],
    "date_max": conn.execute("SELECT MAX(date) FROM fixtures WHERE home_score IS NOT NULL").fetchone()[0],
}
conn.close()

# ---------------- write workbook ----------------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
HEAD = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="2F4858")
TITLE = Font(bold=True, size=14); SUB = Font(italic=True, color="666666")
GREEN = PatternFill("solid", fgColor="C6EFCE"); AMBER = PatternFill("solid", fgColor="FFEB9C"); RED = PatternFill("solid", fgColor="FFC7CE")
THIN = Border(*[Side(style="thin", color="DDDDDD")] * 4)
CEN = Alignment(horizontal="center")

def fillpct(p):
    if p is None: return None
    return GREEN if p >= 72 else (AMBER if p >= 65 else RED)

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.font = HEAD; cell.fill = HFILL; cell.alignment = CEN

def autow(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 1: Overview ---
ws = wb.active; ws.title = "Overview"
ws["A1"] = "OddsFlow V4 — Full-Database Consolidation"; ws["A1"].font = TITLE
ws["A2"] = f"Generated {STAMP} · backtest of every settled+classifiable fixture"; ws["A2"].font = SUB
r = 4
for k, v in [
    ("Fixtures (total)", landscape["fixtures_total"]),
    ("Fixtures settled (have scores)", landscape["settled"]),
    ("Settled + fully classifiable (used in backtest)", landscape["classifiable_settled"]),
    ("Fixture-stats rows (corner data)", landscape["corner_rows"]),
    ("Leagues (total / active)", f"{landscape['leagues']} / {landscape['leagues_active']}"),
    ("Emitted picks logged", landscape["emit_log"]),
    ("Settled date range", f"{landscape['date_min']}  →  {landscape['date_max']}"),
]:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True); ws.cell(row=r, column=2, value=v); r += 1
r += 1
ws.cell(row=r, column=1, value="Method / conventions").font = Font(bold=True, size=12); r += 1
for line in [
    "Partition = (draw_zone × BTS over/under) = 8 cells. Signals (spread, DF, H2H) are NOT axes.",
    "Zones: strong 2.90–3.30 · standard 3.30–3.80 · low 3.80–4.30 · one_sided ≥4.30 · excluded <2.90.",
    "BTS over = btts_yes_odd ≤ btts_no_odd. Spread strong = min(yes,no) < 1.50. DF = |round(home)-round(away)|.",
    "Goals O1.5 = total goals ≥ 2. Corners O7.5 (strong) / O8.5 (else), graded only where corner stats exist.",
    "3-Way = alpha-or-draw: favourite win OR draw = WIN (binary, no void). Hit rate = wins / settled.",
    "Backtest sheets use ALL settled fixtures; 'Live Emits' uses only the picks the engine actually fired.",
    "Green ≥72% · Amber 65–72% · Red <65%.",
]:
    ws.cell(row=r, column=1, value="• " + line); r += 1
autow(ws, [46, 60])

# --- Sheet 2: Cell Matrix ---
ws = wb.create_sheet("Cell Matrix")
hdr = ["Zone", "BTS", "n (fixtures)", "n (corners)", "Goals %", "Goals", "Corners %", "Corners", "3-Way %", "3-Way", "Composite %"]
ws.append(hdr); style_header(ws, 1, len(hdr))
for z in ZONES:
    for b in BTS:
        cl = cells.get((z, b))
        if not cl: continue
        g, c2, t = cl["goals_nl"], cl["corners_nl"], cl["threeway"]
        comp = [x for x in (g.pct(), c2.pct(), t.pct()) if x is not None]
        composite = round(sum(comp) / len(comp), 1) if comp else None
        ws.append([z, b, t.n, corner_avail.get((z, b), 0), g.pct(), g.frac(), c2.pct(), c2.frac(), t.pct(), t.frac(), composite])
        rr = ws.max_row
        for col, p in ((5, g.pct()), (7, c2.pct()), (9, t.pct()), (11, composite)):
            f = fillpct(p)
            if f: ws.cell(row=rr, column=col).fill = f
ws.freeze_panes = "A2"; autow(ws, [11, 8, 12, 12, 9, 9, 10, 9, 9, 9, 12])

# --- Sheet 3: By Tier ---
ws = wb.create_sheet("By Tier")
ws.append(["Tier"] + [MLABEL[m] + " %" for m in MKTS] + [MLABEL[m] for m in MKTS] + ["All markets %"])
style_header(ws, 1, 7)
for t in (1, 2, 3):
    accs = [tier_mkt.get((t, m), Acc()) for m in MKTS]
    tw = sum(a.g for a in accs); tn = sum(a.n for a in accs)
    ws.append([f"T{t}"] + [a.pct() for a in accs] + [a.frac() for a in accs] + [round(100 * tw / tn, 1) if tn else None])
    rr = ws.max_row
    for col, a in zip((2, 3, 4), accs):
        f = fillpct(a.pct());  ws.cell(row=rr, column=col).fill = f if f else PatternFill()
ws.freeze_panes = "A2"; autow(ws, [8, 14, 18, 22, 12, 14, 18, 14])

# --- Sheet 4: By Country ---
ws = wb.create_sheet("By Country")
ws.append(["Country", "n", "Goals %", "Corners %", "3-Way %", "All %"]); style_header(ws, 1, 6)
order = sorted(country_n.items(), key=lambda kv: -kv[1])
for country, n in order:
    accs = [country_mkt.get((country, m), Acc()) for m in MKTS]
    tw = sum(a.g for a in accs); tn = sum(a.n for a in accs)
    ws.append([country, n, accs[0].pct(), accs[1].pct(), accs[2].pct(), round(100 * tw / tn, 1) if tn else None])
    rr = ws.max_row
    for col, a in zip((3, 4, 5), accs):
        f = fillpct(a.pct())
        if f: ws.cell(row=rr, column=col).fill = f
ws.freeze_panes = "A2"; autow(ws, [24, 8, 10, 11, 10, 9])

# --- Sheet 5: Signals ---
ws = wb.create_sheet("Signals")
ws.append(["SPREAD signal (BTS)"]); ws["A1"].font = Font(bold=True, size=12)
ws.append(["Spread"] + [MLABEL[m] + " %" for m in MKTS] + [MLABEL[m] for m in MKTS]); style_header(ws, 2, 7)
for sp in ("strong", "slight"):
    accs = [spread_mkt.get((sp, m), Acc()) for m in MKTS]
    ws.append([sp] + [a.pct() for a in accs] + [a.frac() for a in accs])
    rr = ws.max_row
    for col, a in zip((2, 3, 4), accs):
        f = fillpct(a.pct());
        if f: ws.cell(row=rr, column=col).fill = f
ws.append([])
rdf = ws.max_row + 1
ws.cell(row=rdf, column=1, value="DF signal (draw favouritism)").font = Font(bold=True, size=12)
ws.append(["DF"] + [MLABEL[m] + " %" for m in MKTS] + [MLABEL[m] for m in MKTS]); style_header(ws, rdf + 1, 7)
for d in ("DF0", "DF1", "DF2"):
    accs = [df_mkt.get((d, m), Acc()) for m in MKTS]
    ws.append([d] + [a.pct() for a in accs] + [a.frac() for a in accs])
    rr = ws.max_row
    for col, a in zip((2, 3, 4), accs):
        f = fillpct(a.pct())
        if f: ws.cell(row=rr, column=col).fill = f
autow(ws, [16, 14, 18, 22, 12, 14, 18])

# --- Sheet 6: Rankings ---
ws = wb.create_sheet("Rankings")
ws.append(["Cells ranked by composite hit %"]); ws["A1"].font = Font(bold=True, size=12)
ws.append(["Zone", "BTS", "n", "Composite %", "Goals %", "Corners %", "3-Way %"]); style_header(ws, 2, 7)
ranked = []
for (z, b), cl in cells.items():
    comp = [x for x in (cl["goals_nl"].pct(), cl["corners_nl"].pct(), cl["threeway"].pct()) if x is not None]
    composite = round(sum(comp) / len(comp), 1) if comp else 0
    ranked.append((composite, z, b, cl))
for composite, z, b, cl in sorted(ranked, reverse=True):
    ws.append([z, b, cl["threeway"].n, composite, cl["goals_nl"].pct(), cl["corners_nl"].pct(), cl["threeway"].pct()])
    f = fillpct(composite)
    if f: ws.cell(row=ws.max_row, column=4).fill = f
autow(ws, [11, 8, 8, 13, 9, 10, 9])

# --- Sheet 7: Live Emits ---
ws = wb.create_sheet("Live Emits")
ws.append(["Actual emitted picks — settled on the fly (the engine's real-world record)"]); ws["A1"].font = Font(bold=True, size=12)
ws.append(["Market", "Hit %", "Wins", "Settled"]); style_header(ws, 2, 4)
for m in MKTS:
    a = emit_mkt[m]; ws.append([MLABEL[m], a.pct(), a.g, a.n])
    f = fillpct(a.pct())
    if f: ws.cell(row=ws.max_row, column=2).fill = f
ws.append([])
base = ws.max_row + 1
ws.cell(row=base, column=1, value="Emits by tier × market").font = Font(bold=True)
ws.append(["Tier"] + [MLABEL[m] for m in MKTS]); style_header(ws, base + 1, 4)
for t in (1, 2, 3):
    ws.append([f"T{t}"] + [emit_tier.get((t, m), Acc()).pct() for m in MKTS])
    for col, m in zip((2, 3, 4), MKTS):
        f = fillpct(emit_tier.get((t, m), Acc()).pct())
        if f: ws.cell(row=ws.max_row, column=col).fill = f
autow(ws, [24, 12, 10, 12])

wb.save(XLSX)
print("WROTE", XLSX)

# ---------------- summary txt ----------------
with open(TXT, "w", encoding="utf-8") as fh:
    def w(s): fh.write(s + "\n"); print(s)
    w(f"OddsFlow V4 consolidation — {STAMP}")
    w(f"settled+classifiable fixtures backtested: {len(rows)}")
    w(f"date range: {landscape['date_min']} -> {landscape['date_max']}")
    w("")
    w("CELL MATRIX (composite = mean of 3 market hit%):")
    for z in ZONES:
        for b in BTS:
            cl = cells.get((z, b))
            if not cl: continue
            comp = [x for x in (cl["goals_nl"].pct(), cl["corners_nl"].pct(), cl["threeway"].pct()) if x is not None]
            composite = round(sum(comp) / len(comp), 1) if comp else None
            w(f"  {z:10} {b:6} n={cl['threeway'].n:<6} G {cl['goals_nl'].pct()}  C {cl['corners_nl'].pct()}  3W {cl['threeway'].pct()}  comp {composite}")
    w("")
    w("BY TIER (all markets blended):")
    for t in (1, 2, 3):
        accs = [tier_mkt.get((t, m), Acc()) for m in MKTS]
        tw = sum(a.g for a in accs); tn = sum(a.n for a in accs)
        w(f"  T{t}: {round(100*tw/tn,1) if tn else 0}%  (G {accs[0].pct()} / C {accs[1].pct()} / 3W {accs[2].pct()})")
    w("")
    w("SIGNALS:")
    for sp in ("strong", "slight"):
        w(f"  spread {sp}: G {spread_mkt.get((sp,'goals_nl'),Acc()).pct()}  C {spread_mkt.get((sp,'corners_nl'),Acc()).pct()}  3W {spread_mkt.get((sp,'threeway'),Acc()).pct()}")
    for d in ("DF0", "DF1", "DF2"):
        w(f"  {d}: 3W {df_mkt.get((d,'threeway'),Acc()).pct()}")
    w("")
    w("LIVE EMITS (fired picks):")
    for m in MKTS:
        a = emit_mkt[m]; w(f"  {MLABEL[m]}: {a.pct()}% ({a.g}/{a.n})")
print("WROTE", TXT)
